import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = 'backend/patient_records.xlsx'

# Check if the Excel file exists
if not os.path.exists(EXCEL_FILE):
    # If it doesn't, create a new workbook and save it
    workbook = Workbook()
    sheet = workbook.active
    # Add headers or any initial content if needed
    sheet.append(["Name", "Age", "Gender", "Contact", "Address", "Allergies", "Chronic Conditions"])
    workbook.save(EXCEL_FILE)

# Now load the workbook
workbook = load_workbook(EXCEL_FILE)


# Define a route to handle form submissions
@app.route('/submit_record', methods=['POST'])
def submit_record():
    # Get form data
    name = request.form['name']
    age = request.form['age']
    gender = request.form['gender']
    contact = request.form['contact']
    address = request.form.get('address', '')
    allergies = request.form.get('allergies', '')
    chronic_conditions = request.form.get('chronic_conditions', '')

    # Open the Excel file and append new row
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([name, age, gender, contact, address, allergies, chronic_conditions])
    workbook.save(EXCEL_FILE)

    return "Record added successfully"
